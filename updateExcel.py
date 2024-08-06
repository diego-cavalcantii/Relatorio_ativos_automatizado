import openpyxl
import os
from datetime import datetime

ativos = ["BTG CDB Plus FIRF CrPr","16,10% a.a.", "15,86% a.a.", "12,79% a.a.", "IPCA + 6,20%", "13,58% a.a.","IPCA + 9,46%","IPCA + 9,11%", "BTG PACT CRED CORP II PREV FC FI MULT CP","EXES PREV FIE FC FI MULT CRED PRIV", "Conta corrente"]

def atualizar_celula_com_data():
    def obter_nome_arquivo():
        while True:
            nome_arquivo = input("Digite o nome do arquivo Excel : ") + ".xlsx"
            caminho_arquivo = os.path.join(os.getcwd(), nome_arquivo)
            if os.path.exists(caminho_arquivo):
                print(f"Arquivo '{nome_arquivo}' encontrado.")
                return caminho_arquivo
            else:
                print(f"O arquivo '{nome_arquivo}' não foi encontrado. Por favor, tente novamente.")
    
    caminho_arquivo = obter_nome_arquivo()

    try:
        # Abre o arquivo Excel existente
        workbook = openpyxl.load_workbook(caminho_arquivo)

        # Obtém a planilha chamada 'Dados'
        def obter_planilha(workbook):
            while True:
                nome_planilha = input("Digite o nome da planilha (ex: Dados): ")
                if nome_planilha in workbook.sheetnames:
                    sheet = workbook[nome_planilha]
                    print(f"Planilha '{nome_planilha}' encontrada.")
                    return sheet
                else:
                    print(f"A planilha '{nome_planilha}' não foi encontrada. Por favor, tente novamente.")
        
        sheet = obter_planilha(workbook)

        # Solicitar a posição da primeira célula líquida
        primeira_celula_liquida = input("Digite a posição da primeira célula líquida (ex: Z6): ")

        # Separar a coluna e a linha da célula fornecida
        col, row = openpyxl.utils.cell.coordinate_from_string(primeira_celula_liquida)
        row = int(row)  # Converter a linha para inteiro

        # Atualizar valores na ordem da lista de ativos
        def atualiza_valor_ordenado(ativos, sheet, col, row):
            for i in range(len(ativos)):
                # Calcular a célula líquida e a célula bruta para o ativo atual
                cell_liquida = f"{col}{row + i}"
                cell_bruto = f"{openpyxl.utils.cell.get_column_letter(openpyxl.utils.cell.column_index_from_string(col) + 1)}{row + i}"

                # Pedir os valores do usuário
                valor_liquido = input(f"Digite o valor do {ativos[i]} líquido: ")
                valor_bruto = input(f"Digite o valor do {ativos[i]} bruto: ")

                # Atribuir os valores às células
                sheet[cell_liquida].value = valor_liquido
                sheet[cell_bruto].value = valor_bruto

        atualiza_valor_ordenado(ativos, sheet, col, row)

        # Salva o arquivo Excel com as alterações
        workbook.save(caminho_arquivo)
        print('Informação adicionada com sucesso!')
    except Exception as error:
        print('Erro ao atualizar célula:', error)

# Chama a função para atualizar a célula com a data
atualizar_celula_com_data()
